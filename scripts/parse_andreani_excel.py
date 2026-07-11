#!/usr/bin/env python3
"""Parse Andreani Excel using openpyxl - outputs products as JSON."""

import json
import sys
import openpyxl

def parse_xlsx(path, limit=None):
    products = []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Headers from row 1 (1-indexed in openpyxl)
    headers = [cell.value for cell in ws[1]]
    print(f"Headers ({len(headers)} cols): {headers[:11]}...", file=sys.stderr)

    # Column indices (0-based)
    COL = {
        'referencia': 0,    # A
        'nombre': 1,        # B
        'barcode': 2,       # C
        'tarifeCode': 3,    # D
        'pvp': 4,           # E
        'family': 5,        # F
        'subfamily': 6,     # G
        'descripcion': 7,    # H
        'image': 8,          # I
        'gallery': 9,       # J
        'documents': 10,     # K
        'talla': 11,        # L
        'altura': 12,       # M
        'recorrido': 13,    # N
        'longitud': 14,     # O
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if limit and row_idx - 1 > limit + 1:
            break

        referencia = row[COL['referencia']]
        if not referencia:
            continue

        nombre = row[COL['nombre']] or ''

        # Parse PVP: "18.00 €" or 18.0 or None
        pvp_raw = row[COL['pvp']]
        pvp_cents = 0
        if pvp_raw:
            if isinstance(pvp_raw, (int, float)):
                pvp_cents = int(pvp_raw * 100)
            else:
                try:
                    pvp_str = str(pvp_raw).replace('€', '').replace(',', '.').strip()
                    pvp_cents = int(float(pvp_str) * 100)
                except:
                    pvp_cents = 0

        # Parse image and gallery
        image_url = row[COL['image']] or ''
        gallery_urls = row[COL['gallery']] or ''

        products.append({
            'referencia': str(referencia).strip(),
            'nombre': str(nombre).strip() if nombre else '',
            'barcode': str(row[COL['barcode']] or '').strip(),
            'tarifeCode': str(row[COL['tarifeCode']] or '').strip(),
            'pvpEur': pvp_cents,
            'family': str(row[COL['family']] or '').strip(),
            'subfamily': str(row[COL['subfamily']] or '').strip(),
            'descripcion': str(row[COL['descripcion']] or '').strip(),
            'image': str(image_url).strip(),
            'gallery': str(gallery_urls).strip(),
            'documents': str(row[COL['documents']] or '').strip(),
            'talla': str(row[COL['talla']] or '').strip(),
            'alturaMm': row[COL['altura']],
            'recorridoMm': row[COL['recorrido']],
            'longitudMm': row[COL['longitud']],
        })

    wb.close()
    return products

if __name__ == '__main__':
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else '/tmp/andreani-capture/megatarifas-excel-1781159915299.xlsx'
    limit = int(sys.argv[2]) if len(sys.argv) > 2 else None
    products = parse_xlsx(xlsx_path, limit=limit)
    print(json.dumps(products, ensure_ascii=False))
